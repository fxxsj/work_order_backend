"""
施工单任务导出 Mixin
提供Excel导出功能
"""

from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import logging

from workorder.models.core import WorkOrderTask

logger = logging.getLogger(__name__)


class TaskExportMixin:
    """
    任务导出 Mixin

    提供Excel导出功能，支持筛选条件导出
    """

    @action(detail=False, methods=['post'], url_path='export')
    def export_excel(self, request):
        """
        导出任务列表到Excel

        请求参数：
        - filters: 筛选条件（可选，与list接口相同）
        - task_ids: 指定导出的任务ID列表（可选，优先级高于filters）
        - columns: 指定导出的列（可选，默认导出所有列）

        返回：Excel文件流
        """
        from openpyxl.utils import get_column_letter

        # 获取导出参数
        task_ids = request.data.get('task_ids', [])
        filters = request.data.get('filters', {})
        columns = request.data.get('columns', [
            'id', 'work_order_number', 'process_name', 'task_type',
            'work_content', 'assigned_department', 'assigned_operator',
            'production_quantity', 'quantity_completed', 'progress',
            'priority', 'status', 'created_at', 'updated_at'
        ])

        # 构建查询集
        if task_ids:
            # 优先使用指定任务ID
            queryset = self.get_queryset().filter(id__in=task_ids)
        else:
            # 使用筛选条件
            queryset = self.filter_queryset(self.get_queryset())

        # 限制导出数量（防止导出过多数据）
        max_export = 10000
        if queryset.count() > max_export:
            return Response(
                {'error': f'导出数据量过大（{queryset.count()}条），最多支持导出{max_export}条'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '任务列表'

        # 定义列标题和宽度
        column_config = {
            'id': {'title': 'ID', 'width': 10},
            'work_order_number': {'title': '施工单号', 'width': 18},
            'process_name': {'title': '工序', 'width': 15},
            'task_type': {'title': '任务类型', 'width': 12},
            'work_content': {'title': '任务内容', 'width': 30},
            'assigned_department': {'title': '分派部门', 'width': 15},
            'assigned_operator': {'title': '分派操作员', 'width': 12},
            'production_quantity': {'title': '生产数量', 'width': 12},
            'quantity_completed': {'title': '完成数量', 'width': 12},
            'quantity_defective': {'title': '不良品数量', 'width': 12},
            'progress': {'title': '进度(%)', 'width': 10},
            'priority': {'title': '优先级', 'width': 10},
            'status': {'title': '状态', 'width': 12},
            'created_at': {'title': '创建时间', 'width': 18},
            'updated_at': {'title': '更新时间', 'width': 18},
        }

        # 设置列标题
        header_style = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, col_key in enumerate(columns, 1):
            col_config = column_config.get(col_key, {'title': col_key, 'width': 15})
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_config['title']
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_config['width']

        # 填充数据
        status_display_map = dict(WorkOrderTask.STATUS_CHOICES)
        task_type_display_map = dict(WorkOrderTask.TASK_TYPE_CHOICES)
        priority_display_map = dict(WorkOrderTask.PRIORITY_CHOICES)

        row_idx = 2
        for task in queryset:
            row_data = self._get_row_data(task, columns, status_display_map, task_type_display_map, priority_display_map)

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border

                # 状态列着色
                if columns[col_idx - 1] == 'status':
                    if task.status == 'completed':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif task.status == 'cancelled':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    elif task.status == 'draft':
                        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

            row_idx += 1

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 设置行高
        ws.row_dimensions[1].height = 25

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'任务列表_{timestamp}.xlsx'

        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)

        return response

    def _get_row_data(self, task, columns, status_map, type_map, priority_map):
        """获取单行数据"""
        row = []

        for col_key in columns:
            value = ''
            if col_key == 'id':
                value = task.id
            elif col_key == 'work_order_number':
                value = getattr(task.work_order_process.work_order, 'order_number', '') if task.work_order_process and task.work_order_process.work_order else ''
            elif col_key == 'process_name':
                value = str(task.work_order_process.process) if task.work_order_process and task.work_order_process.process else ''
            elif col_key == 'task_type':
                value = type_map.get(task.task_type, task.task_type) if task.task_type else ''
            elif col_key == 'work_content':
                value = task.work_content or ''
            elif col_key == 'assigned_department':
                value = str(task.assigned_department) if task.assigned_department else ''
            elif col_key == 'assigned_operator':
                value = str(task.assigned_operator) if task.assigned_operator else ''
            elif col_key == 'production_quantity':
                value = task.production_quantity or 0
            elif col_key == 'quantity_completed':
                value = task.quantity_completed or 0
            elif col_key == 'quantity_defective':
                value = task.quantity_defective or 0
            elif col_key == 'progress':
                if task.production_quantity and task.production_quantity > 0:
                    value = round((task.quantity_completed or 0) / task.production_quantity * 100, 1)
                else:
                    value = 0
            elif col_key == 'priority':
                value = priority_map.get(task.priority, task.priority) if task.priority else ''
            elif col_key == 'status':
                value = status_map.get(task.status, task.status) if task.status else ''
            elif col_key == 'created_at':
                value = task.created_at.strftime('%Y-%m-%d %H:%M') if task.created_at else ''
            elif col_key == 'updated_at':
                value = task.updated_at.strftime('%Y-%m-%d %H:%M') if task.updated_at else ''

            row.append(value)

        return row