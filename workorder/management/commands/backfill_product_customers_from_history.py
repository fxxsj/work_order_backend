"""从历史订单表安全回填产品与客户的关联关系。"""

from collections import defaultdict
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from workorder.models.base import Customer
from workorder.models.products import Product, ProductCustomer


DEFAULT_SHEET = "合并查询"
DEFAULT_CUSTOMER_HEADER = "客户"
DEFAULT_PRODUCT_HEADER = "产品名称"
DEFAULT_UNIT_HEADER = "单位"
MAX_UNMATCHED_PREVIEW = 100


def normalize_name(value):
    """统一全半角与空白，保留名称语义用于精确匹配。"""

    text = unicodedata.normalize("NFKC", str(value or ""))
    return " ".join(text.strip().split()).casefold()


class Command(BaseCommand):
    help = (
        "从历史订单 XLSX 的客户和产品名称回填 ProductCustomer；"
        "默认仅预检，使用 --apply 才会写入"
    )

    def add_arguments(self, parser):
        parser.add_argument("file", help="历史订单 XLSX 文件路径")
        parser.add_argument(
            "--sheet",
            default=DEFAULT_SHEET,
            help=f"数据工作表名称（默认：{DEFAULT_SHEET}）",
        )
        parser.add_argument(
            "--customer-header",
            default=DEFAULT_CUSTOMER_HEADER,
            help=f"客户列标题（默认：{DEFAULT_CUSTOMER_HEADER}）",
        )
        parser.add_argument(
            "--product-header",
            default=DEFAULT_PRODUCT_HEADER,
            help=f"产品列标题（默认：{DEFAULT_PRODUCT_HEADER}）",
        )
        parser.add_argument(
            "--unit-header",
            default=DEFAULT_UNIT_HEADER,
            help=(
                f"单位列标题，用于消除同名产品歧义"
                f"（默认：{DEFAULT_UNIT_HEADER}；列不存在时跳过）"
            ),
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="正式创建缺失的产品客户关联；不传时仅预检",
        )
        parser.add_argument(
            "--allow-unmatched",
            action="store_true",
            help="允许正式写入可匹配部分，并跳过无法匹配的名称",
        )
        parser.add_argument(
            "--allow-ambiguous",
            action="store_true",
            help="允许正式写入可匹配部分，并跳过重名歧义关系",
        )
        parser.add_argument(
            "--show-unmatched",
            type=int,
            default=0,
            metavar="N",
            help="显示最多 N 个未匹配/歧义名称（默认不显示，最大 100）",
        )
        parser.add_argument(
            "--max-rows",
            type=int,
            default=200_000,
            help="允许扫描的最大数据行数（默认 200000）",
        )
        parser.add_argument(
            "--max-file-size-mb",
            type=int,
            default=100,
            help="允许的最大文件大小，单位 MB（默认 100）",
        )

    def handle(self, *args, **options):
        file_path = self._validate_options(options)
        source = self._read_source_pairs(
            file_path=file_path,
            sheet_name=options["sheet"],
            customer_header=options["customer_header"],
            product_header=options["product_header"],
            unit_header=options["unit_header"],
            max_rows=options["max_rows"],
        )
        report = self._match_pairs(source)
        self._print_report(report, source, options["show_unmatched"])

        if not options["apply"]:
            self.stdout.write(
                self.style.WARNING(
                    "\n当前为预检模式，数据库未发生变化；确认结果后添加 --apply"
                )
            )
            return

        if report["ambiguous_pairs"] and not options["allow_ambiguous"]:
            raise CommandError(
                "存在重名歧义，已拒绝写入。确认只写入无歧义部分后使用 "
                "--apply --allow-ambiguous。"
            )
        if report["unmatched_pairs"] and not options["allow_unmatched"]:
            raise CommandError(
                "存在未匹配记录，已拒绝写入。确认可跳过后使用 "
                "--apply --allow-unmatched。"
            )
        if not report["matched_pairs"]:
            raise CommandError("没有可写入的匹配关系，已拒绝执行。")

        with transaction.atomic():
            ProductCustomer.objects.bulk_create(
                [
                    ProductCustomer(product_id=product_id, customer_id=customer_id)
                    for product_id, customer_id in report["to_create_pairs"]
                ],
                batch_size=1000,
                ignore_conflicts=True,
            )

        created_count = len(report["to_create_pairs"])
        self.stdout.write(
            self.style.SUCCESS(
                f"\n回填完成：新增 {created_count} 条关联，"
                f"保留 {len(report['existing_pairs'])} 条已有关系。"
            )
        )

    def _validate_options(self, options):
        file_path = Path(options["file"]).expanduser().resolve()
        if not file_path.is_file():
            raise CommandError(f"文件不存在：{file_path}")
        if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise CommandError("仅支持 .xlsx 或 .xlsm 文件。")
        if options["max_rows"] <= 0:
            raise CommandError("--max-rows 必须大于 0。")
        if options["max_file_size_mb"] <= 0:
            raise CommandError("--max-file-size-mb 必须大于 0。")
        if not 0 <= options["show_unmatched"] <= MAX_UNMATCHED_PREVIEW:
            raise CommandError(
                f"--show-unmatched 必须在 0 到 {MAX_UNMATCHED_PREVIEW} 之间。"
            )

        max_bytes = options["max_file_size_mb"] * 1024 * 1024
        if file_path.stat().st_size > max_bytes:
            raise CommandError(f"文件超过 {options['max_file_size_mb']} MB 限制。")
        return file_path

    def _read_source_pairs(
        self,
        *,
        file_path,
        sheet_name,
        customer_header,
        product_header,
        unit_header,
        max_rows,
    ):
        try:
            workbook = load_workbook(
                file_path,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise CommandError(f"无法读取工作簿：{exc}") from exc

        try:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(
                    f"工作表不存在：{sheet_name}；"
                    f"可用工作表：{', '.join(workbook.sheetnames)}"
                )
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration as exc:
                raise CommandError("工作表为空。") from exc

            normalized_headers = {
                normalize_name(value): index
                for index, value in enumerate(headers)
                if normalize_name(value)
            }
            customer_key = normalize_name(customer_header)
            product_key = normalize_name(product_header)
            if customer_key not in normalized_headers:
                raise CommandError(f"缺少客户列：{customer_header}")
            if product_key not in normalized_headers:
                raise CommandError(f"缺少产品列：{product_header}")

            customer_index = normalized_headers[customer_key]
            product_index = normalized_headers[product_key]
            unit_index = normalized_headers.get(normalize_name(unit_header))
            pairs = defaultdict(set)
            labels = {"customer": {}, "product": {}}
            scanned_rows = 0
            blank_rows = 0

            for row_number, row in enumerate(rows, start=2):
                scanned_rows += 1
                if scanned_rows > max_rows:
                    raise CommandError(f"数据行超过 --max-rows={max_rows} 限制")
                customer_value = (
                    row[customer_index] if customer_index < len(row) else None
                )
                product_value = row[product_index] if product_index < len(row) else None
                unit_value = (
                    row[unit_index]
                    if unit_index is not None and unit_index < len(row)
                    else None
                )
                customer_name = normalize_name(customer_value)
                product_name = normalize_name(product_value)
                unit_name = normalize_name(unit_value)
                if not customer_name or not product_name:
                    blank_rows += 1
                    continue

                if unit_name:
                    pairs[(customer_name, product_name)].add(unit_name)
                else:
                    pairs[(customer_name, product_name)]
                labels["customer"].setdefault(
                    customer_name, str(customer_value).strip()
                )
                labels["product"].setdefault(product_name, str(product_value).strip())

            return {
                "pairs": dict(pairs),
                "labels": labels,
                "scanned_rows": scanned_rows,
                "blank_rows": blank_rows,
            }
        finally:
            workbook.close()

    def _match_pairs(self, source):
        customer_index = self._build_name_index(
            Customer.objects.values_list("id", "name")
        )
        product_index = self._build_product_index(
            Product.objects.values_list("id", "name", "unit")
        )

        matched_pairs = set()
        unmatched_pairs = set()
        ambiguous_pairs = set()
        unmatched_customers = set()
        unmatched_products = set()
        ambiguous_customers = set()
        ambiguous_products = set()
        resolved_by_unit = set()

        for (customer_name, product_name), source_units in source["pairs"].items():
            customer_ids = customer_index.get(customer_name, ())
            product_candidates = product_index.get(product_name, ())
            product_ids = tuple(candidate[0] for candidate in product_candidates)

            if len(product_candidates) > 1 and source_units:
                unit_matched_ids = tuple(
                    product_id
                    for product_id, product_unit in product_candidates
                    if product_unit and product_unit in source_units
                )
                if unit_matched_ids:
                    product_ids = unit_matched_ids
                if len(unit_matched_ids) == 1:
                    resolved_by_unit.add((customer_name, product_name))

            if not customer_ids:
                unmatched_customers.add(customer_name)
            elif len(customer_ids) > 1:
                ambiguous_customers.add(customer_name)
            if not product_ids:
                unmatched_products.add(product_name)
            elif len(product_ids) > 1:
                ambiguous_products.add(product_name)

            if len(customer_ids) == 1 and len(product_ids) == 1:
                matched_pairs.add((product_ids[0], customer_ids[0]))
            elif len(customer_ids) > 1 or len(product_ids) > 1:
                ambiguous_pairs.add((customer_name, product_name))
            else:
                unmatched_pairs.add((customer_name, product_name))

        product_ids = {pair[0] for pair in matched_pairs}
        customer_ids = {pair[1] for pair in matched_pairs}
        existing_pairs = set()
        if product_ids and customer_ids:
            existing_pairs = set(
                ProductCustomer.objects.filter(
                    product_id__in=product_ids,
                    customer_id__in=customer_ids,
                ).values_list("product_id", "customer_id")
            )
            existing_pairs &= matched_pairs

        return {
            "matched_pairs": matched_pairs,
            "unmatched_pairs": unmatched_pairs,
            "ambiguous_pairs": ambiguous_pairs,
            "unmatched_customers": unmatched_customers,
            "unmatched_products": unmatched_products,
            "ambiguous_customers": ambiguous_customers,
            "ambiguous_products": ambiguous_products,
            "resolved_by_unit": resolved_by_unit,
            "existing_pairs": existing_pairs,
            "to_create_pairs": matched_pairs - existing_pairs,
        }

    @staticmethod
    def _build_name_index(rows):
        index = defaultdict(list)
        for object_id, name in rows:
            normalized = normalize_name(name)
            if normalized:
                index[normalized].append(object_id)
        return {name: tuple(ids) for name, ids in index.items()}

    @staticmethod
    def _build_product_index(rows):
        index = defaultdict(list)
        for object_id, name, unit in rows:
            normalized = normalize_name(name)
            if normalized:
                index[normalized].append((object_id, normalize_name(unit)))
        return {name: tuple(candidates) for name, candidates in index.items()}

    def _print_report(self, report, source, show_unmatched):
        self.stdout.write("产品—客户历史关系回填预检")
        self.stdout.write(f"扫描数据行：{source['scanned_rows']}")
        self.stdout.write(f"跳过空值行：{source['blank_rows']}")
        self.stdout.write(f"源文件唯一关系：{len(source['pairs'])}")
        self.stdout.write(f"可精确匹配关系：{len(report['matched_pairs'])}")
        self.stdout.write(f"其中按单位消除同名歧义：{len(report['resolved_by_unit'])}")
        self.stdout.write(f"已有关系：{len(report['existing_pairs'])}")
        self.stdout.write(f"待新增关系：{len(report['to_create_pairs'])}")
        self.stdout.write(f"未匹配关系：{len(report['unmatched_pairs'])}")
        self.stdout.write(f"未匹配客户名称：{len(report['unmatched_customers'])}")
        self.stdout.write(f"未匹配产品名称：{len(report['unmatched_products'])}")
        self.stdout.write(f"歧义关系：{len(report['ambiguous_pairs'])}")
        self.stdout.write(f"重名客户名称：{len(report['ambiguous_customers'])}")
        self.stdout.write(f"重名产品名称：{len(report['ambiguous_products'])}")

        if not show_unmatched:
            return
        self._print_name_preview(
            "未匹配客户",
            report["unmatched_customers"],
            source["labels"]["customer"],
            show_unmatched,
        )
        self._print_name_preview(
            "未匹配产品",
            report["unmatched_products"],
            source["labels"]["product"],
            show_unmatched,
        )
        self._print_name_preview(
            "重名客户",
            report["ambiguous_customers"],
            source["labels"]["customer"],
            show_unmatched,
        )
        self._print_name_preview(
            "重名产品",
            report["ambiguous_products"],
            source["labels"]["product"],
            show_unmatched,
        )

    def _print_name_preview(self, title, names, labels, limit):
        if not names:
            return
        self.stdout.write(f"\n{title}（最多显示 {limit} 个）：")
        for name in sorted(names)[:limit]:
            self.stdout.write(f"  - {labels.get(name, name)}")
