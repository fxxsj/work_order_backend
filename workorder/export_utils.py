"""
数据导出工具
支持 Excel 格式导出
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status


def create_excel_response(filename):
    """创建 Excel HTTP 响应"""
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def style_header_cell(cell):
    """设置表头单元格样式"""
    if not OPENPYXL_AVAILABLE:
        return
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def style_data_cell(cell):
    """设置数据单元格样式"""
    if not OPENPYXL_AVAILABLE:
        return
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def export_work_orders(queryset, filename=None):
    """
    导出施工单列表到 Excel

    Args:
        queryset: 施工单查询集
        filename: 文件名（可选）

    Returns:
        HttpResponse: Excel 文件响应
    """
    if not OPENPYXL_AVAILABLE:
        from django.http import HttpResponse

        return HttpResponse(
            "Excel 导出功能需要安装 openpyxl 库。请运行: pip install openpyxl",
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content_type="text/plain; charset=utf-8",
        )

    if filename is None:
        filename = (
            f'施工单列表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "施工单列表"

    # 表头
    headers = [
        "施工单号",
        "客户名称",
        "业务员",
        "创建人",
        "创建时间",
        "订单日期",
        "交货日期",
        "状态",
        "审核状态",
        "优先级",
        "备注",
    ]

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        style_header_cell(cell)

    # 状态映射
    status_map = {
        "pending": "待开始",
        "in_progress": "进行中",
        "paused": "已暂停",
        "completed": "已完成",
        "cancelled": "已取消",
    }

    approval_status_map = {
        "pending": "待审核",
        "approved": "已审核",
        "rejected": "已拒绝",
    }

    priority_map = {
        "low": "低",
        "normal": "普通",
        "high": "高",
        "urgent": "紧急",
    }

    # 写入数据
    for row_num, work_order in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=work_order.order_number)
        ws.cell(
            row=row_num,
            column=2,
            value=work_order.customer.name if work_order.customer else "",
        )
        ws.cell(
            row=row_num,
            column=3,
            value=(
                work_order.customer.salesperson.username
                if work_order.customer and work_order.customer.salesperson
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=4,
            value=(
                work_order.created_by.username if work_order.created_by else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=5,
            value=(
                work_order.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if work_order.created_at
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=6,
            value=(
                work_order.order_date.strftime("%Y-%m-%d")
                if work_order.order_date
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=7,
            value=(
                work_order.delivery_date.strftime("%Y-%m-%d")
                if work_order.delivery_date
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=8,
            value=status_map.get(work_order.status, work_order.status),
        )
        ws.cell(
            row=row_num,
            column=9,
            value=approval_status_map.get(
                work_order.approval_status, work_order.approval_status
            ),
        )
        ws.cell(
            row=row_num,
            column=10,
            value=priority_map.get(work_order.priority, work_order.priority),
        )
        ws.cell(row=row_num, column=11, value=work_order.notes or "")

        # 设置数据单元格样式
        for col_num in range(1, len(headers) + 1):
            style_data_cell(ws.cell(row=row_num, column=col_num))

    # 调整列宽
    column_widths = [18, 20, 12, 12, 20, 12, 12, 10, 10, 10, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 创建响应
    response = create_excel_response(filename)
    wb.save(response)
    return response


def export_tasks(queryset, filename=None):
    """
    导出任务列表到 Excel

    Args:
        queryset: 任务查询集
        filename: 文件名（可选）

    Returns:
        HttpResponse: Excel 文件响应
    """
    if not OPENPYXL_AVAILABLE:
        from django.http import HttpResponse

        return HttpResponse(
            "Excel 导出功能需要安装 openpyxl 库。请运行: pip install openpyxl",
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content_type="text/plain; charset=utf-8",
        )

    if filename is None:
        filename = f'任务列表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "任务列表"

    # 表头
    headers = [
        "施工单号",
        "工序",
        "任务类型",
        "工作内容",
        "分派部门",
        "分派操作员",
        "生产数量",
        "完成数量",
        "不良品数量",
        "状态",
        "创建时间",
        "更新时间",
        "备注",
    ]

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        style_header_cell(cell)

    # 状态映射
    status_map = {
        "pending": "待开始",
        "in_progress": "进行中",
        "completed": "已完成",
        "cancelled": "已取消",
        "skipped": "已跳过",
    }

    task_type_map = {
        "general": "通用",
        "artwork": "制版",
        "cutting": "开料",
        "printing": "印刷",
        "foiling": "烫金",
        "embossing": "压凸",
        "die_cutting": "模切",
        "packaging": "包装",
    }

    # 写入数据
    for row_num, task in enumerate(queryset, 2):
        work_order = task.work_order_process.work_order
        process = task.work_order_process.process

        ws.cell(row=row_num, column=1, value=work_order.order_number)
        ws.cell(row=row_num, column=2, value=process.name if process else "")
        ws.cell(
            row=row_num,
            column=3,
            value=task_type_map.get(task.task_type, task.task_type),
        )
        ws.cell(row=row_num, column=4, value=task.work_content or "")
        ws.cell(
            row=row_num,
            column=5,
            value=(
                task.assigned_department.name
                if task.assigned_department
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=6,
            value=(
                task.assigned_operator.username
                if task.assigned_operator
                else ""
            ),
        )
        ws.cell(row=row_num, column=7, value=task.production_quantity or "")
        ws.cell(row=row_num, column=8, value=task.quantity_completed or 0)
        ws.cell(row=row_num, column=9, value=task.quantity_defective or 0)
        ws.cell(
            row=row_num,
            column=10,
            value=status_map.get(task.status, task.status),
        )
        ws.cell(
            row=row_num,
            column=11,
            value=(
                task.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if task.created_at
                else ""
            ),
        )
        ws.cell(
            row=row_num,
            column=12,
            value=(
                task.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                if task.updated_at
                else ""
            ),
        )
        ws.cell(
            row=row_num, column=13, value=task.production_requirements or ""
        )

        # 设置数据单元格样式
        for col_num in range(1, len(headers) + 1):
            style_data_cell(ws.cell(row=row_num, column=col_num))

    # 调整列宽
    column_widths = [18, 15, 10, 30, 15, 12, 12, 12, 12, 10, 20, 20, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 创建响应
    response = create_excel_response(filename)
    wb.save(response)
    return response


def export_customers(queryset, filename=None):
    """
    导出客户列表到 Excel

    Args:
        queryset: 客户查询集
        filename: 文件名（可选）

    Returns:
        HttpResponse: Excel 文件响应
    """
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(
            "Excel 导出功能需要安装 openpyxl 库。请运行: pip install openpyxl",
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content_type="text/plain; charset=utf-8",
        )

    if filename is None:
        filename = f'客户列表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "客户列表"

    # 表头
    headers = [
        "名称",
        "联系人",
        "电话",
        "邮箱",
        "地址",
        "业务员",
        "备注",
        "创建时间",
    ]

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        style_header_cell(cell)

    # 写入数据
    for row_num, customer in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=customer.name or "")
        ws.cell(row=row_num, column=2, value=customer.contact_person or "")
        ws.cell(row=row_num, column=3, value=customer.phone or "")
        ws.cell(row=row_num, column=4, value=customer.email or "")
        ws.cell(row=row_num, column=5, value=customer.address or "")
        ws.cell(
            row=row_num,
            column=6,
            value=(
                customer.salesperson.username if customer.salesperson else ""
            ),
        )
        ws.cell(row=row_num, column=7, value=customer.notes or "")
        ws.cell(
            row=row_num,
            column=8,
            value=(
                customer.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if customer.created_at
                else ""
            ),
        )

        # 设置数据单元格样式
        for col_num in range(1, len(headers) + 1):
            style_data_cell(ws.cell(row=row_num, column=col_num))

    # 调整列宽
    column_widths = [20, 12, 15, 25, 30, 12, 30, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 创建响应
    response = create_excel_response(filename)
    wb.save(response)
    return response


def export_products(queryset, filename=None):
    """
    导出产品列表到 Excel

    Args:
        queryset: 产品查询集
        filename: 文件名（可选）

    Returns:
        HttpResponse: Excel 文件响应
    """
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(
            "Excel 导出功能需要安装 openpyxl 库。请运行: pip install openpyxl",
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content_type="text/plain; charset=utf-8",
        )

    if filename is None:
        filename = f'产品列表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "产品列表"

    # 表头
    headers = [
        "编码",
        "名称",
        "规格",
        "单位",
        "单价",
        "库存数量",
        "最小库存",
        "产品类型",
        "产品组",
        "描述",
        "是否启用",
        "创建时间",
    ]

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        style_header_cell(cell)

    # 产品类型映射
    product_type_map = {
        "single": "单品",
        "group_main": "套装主产品",
        "group_item": "套装子产品",
    }

    # 写入数据
    for row_num, product in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=product.code or "")
        ws.cell(row=row_num, column=2, value=product.name or "")
        ws.cell(row=row_num, column=3, value=product.specification or "")
        ws.cell(row=row_num, column=4, value=product.unit or "")
        ws.cell(
            row=row_num,
            column=5,
            value=str(product.unit_price) if product.unit_price else "0",
        )
        ws.cell(row=row_num, column=6, value=product.stock_quantity or 0)
        ws.cell(row=row_num, column=7, value=product.min_stock_quantity or 0)
        ws.cell(
            row=row_num,
            column=8,
            value=product_type_map.get(
                product.product_type, product.product_type or ""
            ),
        )
        ws.cell(
            row=row_num,
            column=9,
            value=product.product_group.name if product.product_group else "",
        )
        ws.cell(row=row_num, column=10, value=product.description or "")
        ws.cell(
            row=row_num, column=11, value="是" if product.is_active else "否"
        )
        ws.cell(
            row=row_num,
            column=12,
            value=(
                product.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if product.created_at
                else ""
            ),
        )

        # 设置数据单元格样式
        for col_num in range(1, len(headers) + 1):
            style_data_cell(ws.cell(row=row_num, column=col_num))

    # 调整列宽
    column_widths = [15, 20, 20, 8, 10, 10, 10, 12, 15, 30, 10, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 创建响应
    response = create_excel_response(filename)
    wb.save(response)
    return response


def import_products(file, user=None):
    """
    从 Excel 导入产品数据

    Args:
        file: 上传的 Excel 文件
        user: 当前用户（用于审计）

    Returns:
        dict: 包含 success_count, created_count, updated_count,
            error_count, errors 的字典
    """
    if not OPENPYXL_AVAILABLE:
        return {
            "success_count": 0,
            "error_count": 1,
            "errors": [
                "Excel 导入功能需要安装 openpyxl 库。请运行: pip install openpyxl"
            ],
        }

    import io
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from workorder.models.products import Product, ProductGroup

    try:
        file_content = file.read()
        if not file_content:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": ["文件内容为空"],
            }
        try:
            wb = load_workbook(io.BytesIO(file_content))
        except Exception:
            wb = Workbook(io.BytesIO(file_content))
        ws = wb.active
        if ws is None:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": ["Excel 文件格式不正确，无法读取工作表"],
            }

        # 读取表头（第一行）
        if ws.max_row < 1:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": ["Excel 文件为空或格式不正确"],
            }
        first_row = ws[1]
        if first_row is None:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": ["Excel 文件无法读取第一行"],
            }
        headers = [cell.value for cell in first_row]
        # 找到各列索引
        col_map = {}
        for idx, h in enumerate(headers, 1):
            if h:
                h_lower = str(h).strip().lower()
                if "编码" in h_lower or "code" in h_lower:
                    col_map["code"] = idx
                elif "名称" in h_lower and "产品" in h_lower:
                    col_map["name"] = idx
                elif "名称" in h_lower:
                    col_map["name"] = idx
                elif "规格" in h_lower or "spec" in h_lower:
                    col_map["specification"] = idx
                elif "单位" in h_lower or "unit" in h_lower:
                    col_map["unit"] = idx
                elif "单价" in h_lower or "price" in h_lower:
                    col_map["unit_price"] = idx
                elif "库存数量" in h_lower or "stock" in h_lower:
                    col_map["stock_quantity"] = idx
                elif "最小库存" in h_lower or "min_stock" in h_lower:
                    col_map["min_stock_quantity"] = idx
                elif "产品类型" in h_lower or "type" in h_lower:
                    col_map["product_type"] = idx
                elif "产品组" in h_lower or "group" in h_lower:
                    col_map["product_group"] = idx
                elif "描述" in h_lower or "desc" in h_lower:
                    col_map["description"] = idx
                elif "启用" in h_lower or "active" in h_lower:
                    col_map["is_active"] = idx

        # 产品类型映射
        product_type_map = {
            "单品": "single",
            "套装主产品": "group_main",
            "套装子产品": "group_item",
            "single": "single",
            "group_main": "group_main",
            "group_item": "group_item",
        }

        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        imported_codes = set()  # 跟踪本轮导入的产品编码，用于检测重复行

        # 从第二行开始读取数据
        row_iterator = ws.iter_rows(min_row=2, values_only=True)
        for row_num, row in enumerate(row_iterator, 2):
            try:
                if row is None:
                    continue
                data = {}

                code_idx = col_map.get("code")
                if code_idx and code_idx <= len(row):
                    val = row[code_idx - 1]
                    data["code"] = str(val).strip() if val else ""

                name_idx = col_map.get("name")
                if name_idx and name_idx <= len(row):
                    val = row[name_idx - 1]
                    data["name"] = str(val).strip() if val else ""

                spec_idx = col_map.get("specification")
                if spec_idx and spec_idx <= len(row):
                    val = row[spec_idx - 1]
                    data["specification"] = str(val).strip() if val else ""

                unit_idx = col_map.get("unit")
                if unit_idx and unit_idx <= len(row):
                    val = row[unit_idx - 1]
                    data["unit"] = str(val).strip() if val else ""

                price_idx = col_map.get("unit_price")
                if price_idx and price_idx <= len(row):
                    val = row[price_idx - 1]
                    try:
                        data["unit_price"] = float(val) if val else 0
                    except (ValueError, TypeError):
                        data["unit_price"] = 0

                stock_idx = col_map.get("stock_quantity")
                if stock_idx and stock_idx <= len(row):
                    val = row[stock_idx - 1]
                    try:
                        data["stock_quantity"] = int(val) if val else 0
                    except (ValueError, TypeError):
                        data["stock_quantity"] = 0

                min_stock_idx = col_map.get("min_stock_quantity")
                if min_stock_idx and min_stock_idx <= len(row):
                    val = row[min_stock_idx - 1]
                    try:
                        data["min_stock_quantity"] = int(val) if val else 0
                    except (ValueError, TypeError):
                        data["min_stock_quantity"] = 0

                type_idx = col_map.get("product_type")
                if type_idx and type_idx <= len(row):
                    val = row[type_idx - 1]
                    type_str = str(val).strip() if val else ""
                    data["product_type"] = product_type_map.get(
                        type_str, "single"
                    )

                group_idx = col_map.get("product_group")
                if group_idx and group_idx <= len(row):
                    val = row[group_idx - 1]
                    group_name = str(val).strip() if val else ""
                    if group_name:
                        data["product_group_name"] = group_name

                desc_idx = col_map.get("description")
                if desc_idx and desc_idx <= len(row):
                    val = row[desc_idx - 1]
                    data["description"] = str(val).strip() if val else ""

                active_idx = col_map.get("is_active")
                if active_idx and active_idx <= len(row):
                    val = row[active_idx - 1]
                    data["is_active"] = str(val).strip() in (
                        "是",
                        "yes",
                        "true",
                        "1",
                        "启用",
                    )

                if not data.get("code"):
                    errors.append(f"第{row_num}行: 产品编码不能为空")
                    error_count += 1
                    continue

                if not data.get("name"):
                    errors.append(f"第{row_num}行: 产品名称不能为空")
                    error_count += 1
                    continue

                # 检查本批次内是否有重复编码
                code_lower = data["code"].lower()
                if code_lower in imported_codes:
                    errors.append(
                        f'第{row_num}行: 产品编码 "{data["code"]}" 在本次导入中已出现，将跳过'
                    )
                    error_count += 1
                    continue
                imported_codes.add(code_lower)

                # 按编码查找是否已存在
                existing = Product.objects.filter(
                    code__iexact=data["code"]
                ).first()

                if existing:
                    # 更新现有产品
                    for field in [
                        "name",
                        "specification",
                        "unit",
                        "unit_price",
                        "stock_quantity",
                        "min_stock_quantity",
                        "product_type",
                        "description",
                        "is_active",
                    ]:
                        if field in data:
                            setattr(existing, field, data[field])
                    # 处理产品组
                    if "product_group_name" in data:
                        group_name = data["product_group_name"]
                        if group_name:
                            group = ProductGroup.objects.filter(
                                name=group_name
                            ).first()
                            existing.product_group = group
                    existing.save()
                    updated_count += 1
                else:
                    # 创建新产品
                    Product.objects.create(
                        code=data["code"],
                        name=data["name"],
                        specification=data.get("specification", ""),
                        unit=data.get("unit", "件"),
                        unit_price=data.get("unit_price", 0),
                        stock_quantity=data.get("stock_quantity", 0),
                        min_stock_quantity=data.get("min_stock_quantity", 0),
                        product_type=data.get("product_type", "single"),
                        description=data.get("description", ""),
                        is_active=data.get("is_active", True),
                    )
                    created_count += 1

            except Exception as e:
                errors.append(f"第{row_num}行: {str(e)}")
                error_count += 1

        return {
            "success_count": created_count + updated_count,
            "created_count": created_count,
            "updated_count": updated_count,
            "error_count": error_count,
            "errors": errors[:50],  # 最多返回50条错误
        }

    except InvalidFileException as e:
        return {
            "success_count": 0,
            "error_count": 1,
            "errors": [f"无效的 Excel 文件: {str(e)}"],
        }
    except Exception as e:
        return {
            "success_count": 0,
            "error_count": 1,
            "errors": [f"文件读取失败: {str(e)}"],
        }


def import_customers(file, user=None):
    """
    从 Excel 导入客户数据

    Args:
        file: 上传的 Excel 文件
        user: 当前用户（用于业务员分配）

    Returns:
        dict: 包含 success_count, error_count, errors 的字典
    """
    if not OPENPYXL_AVAILABLE:
        return {
            "success_count": 0,
            "error_count": 1,
            "errors": [
                "Excel 导入功能需要安装 openpyxl 库。请运行: pip install openpyxl"
            ],
        }

    import io
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from workorder.models.base import Customer

    try:
        file_content = file.read()
        if not file_content:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": ["文件内容为空"],
            }
        try:
            wb = load_workbook(io.BytesIO(file_content))
        except Exception:
            wb = Workbook(io.BytesIO(file_content))
        ws = wb.active
        if ws is None:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": ["Excel 文件格式不正确，无法读取工作表"],
            }

        # 读取表头（第一行）
        if ws.max_row < 1:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": ["Excel 文件为空或格式不正确"],
            }
        first_row = ws[1]
        if first_row is None:
            return {
                "success_count": 0,
                "error_count": 1,
                "errors": ["Excel 文件无法读取第一行"],
            }
        headers = [cell.value for cell in first_row]
        # 找到各列索引
        col_map = {}
        for idx, h in enumerate(headers, 1):
            if h:
                h_lower = str(h).strip().lower()
                if "名称" in h_lower or "name" in h_lower:
                    col_map["name"] = idx
                elif "联系人" in h_lower or "contact" in h_lower:
                    col_map["contact_person"] = idx
                elif "电话" in h_lower or "phone" in h_lower:
                    col_map["phone"] = idx
                elif "邮箱" in h_lower or "email" in h_lower:
                    col_map["email"] = idx
                elif "地址" in h_lower or "address" in h_lower:
                    col_map["address"] = idx
                elif "备注" in h_lower or "notes" in h_lower:
                    col_map["notes"] = idx

        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        imported_names = set()  # 跟踪本轮导入的客户名称，用于检测重复行

        # 从第二行开始读取数据
        row_iterator = ws.iter_rows(min_row=2, values_only=True)
        for row_num, row in enumerate(row_iterator, 2):
            try:
                if row is None:
                    continue
                data = {}
                name_idx = col_map.get("name")
                if name_idx and name_idx <= len(row):
                    val = row[name_idx - 1]
                    data["name"] = str(val).strip() if val else ""
                contact_idx = col_map.get("contact_person")
                if contact_idx and contact_idx <= len(row):
                    val = row[contact_idx - 1]
                    data["contact_person"] = str(val).strip() if val else ""
                phone_idx = col_map.get("phone")
                if phone_idx and phone_idx <= len(row):
                    val = row[phone_idx - 1]
                    data["phone"] = str(val).strip() if val else ""
                email_idx = col_map.get("email")
                if email_idx and email_idx <= len(row):
                    val = row[email_idx - 1]
                    data["email"] = str(val).strip() if val else ""
                address_idx = col_map.get("address")
                if address_idx and address_idx <= len(row):
                    val = row[address_idx - 1]
                    data["address"] = str(val).strip() if val else ""
                notes_idx = col_map.get("notes")
                if notes_idx and notes_idx <= len(row):
                    val = row[notes_idx - 1]
                    data["notes"] = str(val).strip() if val else ""

                if not data.get("name"):
                    errors.append(f"第{row_num}行: 客户名称不能为空")
                    error_count += 1
                    continue

                # 检查本批次内是否有重复名称
                name_lower = data["name"].lower()
                if name_lower in imported_names:
                    errors.append(
                        f'第{row_num}行: 客户名称 "{data["name"]}" 在本次导入中已出现，将跳过'
                    )
                    error_count += 1
                    continue
                imported_names.add(name_lower)

                # 按名称查找是否已存在（不区分大小写）
                existing = Customer.objects.filter(
                    name__iexact=data["name"]
                ).first()

                if existing:
                    # 更新现有客户
                    for field in [
                        "contact_person",
                        "phone",
                        "email",
                        "address",
                        "notes",
                    ]:
                        if field in data:
                            setattr(existing, field, data[field])
                    existing.save()
                    updated_count += 1
                else:
                    # 创建新客户
                    Customer.objects.create(
                        name=data["name"],
                        contact_person=data.get("contact_person", ""),
                        phone=data.get("phone", ""),
                        email=data.get("email", ""),
                        address=data.get("address", ""),
                        notes=data.get("notes", ""),
                        salesperson=(
                            user if user and user.is_authenticated else None
                        ),
                    )
                    created_count += 1

            except Exception as e:
                errors.append(f"第{row_num}行: {str(e)}")
                error_count += 1

        return {
            "success_count": created_count + updated_count,
            "created_count": created_count,
            "updated_count": updated_count,
            "error_count": error_count,
            "errors": errors[:50],  # 最多返回50条错误
        }

    except InvalidFileException as e:
        return {
            "success_count": 0,
            "error_count": 1,
            "errors": [f"无效的 Excel 文件: {str(e)}"],
        }
    except Exception as e:
        return {
            "success_count": 0,
            "error_count": 1,
            "errors": [f"文件读取失败: {str(e)}"],
        }
