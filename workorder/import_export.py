"""
通用导入/导出框架

提供通用的 Excel 导入/导出功能，支持通过配置定义字段映射。

使用示例:
    # 定义导出配置
    EXPORT_CUSTOMERS = ExportConfig(
        filename='customers',
        sheet_title='客户列表',
        fields=[
            ('名称', 'name', lambda x: x.name),
            ('联系人', 'contact_person', lambda x: x.contact_person),
            ...
        ],
        column_widths=[20, 12, 15],
    )

    # 定义导入配置
    IMPORT_CUSTOMERS = ImportConfig(
        model=Customer,
        unique_field='name',
        field_mappings=[
            ('名称', 'name', str, lambda v: v.strip()),
            ('电话', 'phone', str, lambda v: v.strip()),
            ...
        ],
        update_fields=['contact_person', 'phone', 'email', 'address', 'notes'],
        create_defaults={'salesperson': lambda user: user},
    )

    # 使用
    export_model(queryset, EXPORT_CUSTOMERS)
    import_model(file, IMPORT_CUSTOMERS, user)
"""

from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional, Tuple, Type
from django.db import models


@dataclass
class ExportField:
    """导出字段配置"""
    header: str  # Excel表头名称
    getter: Callable[[Any], Any]  # 从对象获取值的函数


@dataclass
class ExportConfig:
    """导出配置"""
    filename: str  # 文件名前缀
    sheet_title: str  # 工作表标题
    fields: List[ExportField]  # 字段列表
    column_widths: Optional[List[int]] = None  # 列宽列表


@dataclass
class ImportField:
    """导入字段配置"""
    excel_headers: List[str]  # 匹配的Excel表头名称（支持多个别名）
    model_field: str  # 模型字段名
    transformer: Callable[[Any], Any] = lambda x: x  # 值转换函数


@dataclass
class ImportConfig:
    """导入配置"""
    model: Type[models.Model]  # Django模型类
    unique_field: str  # 唯一标识字段（用于upsert）
    field_mappings: List[ImportField]  # 字段映射列表
    update_fields: List[str] = field(default_factory=list)  # 可更新的字段列表
    create_defaults: Dict[str, Any] = field(default_factory=dict)  # 创建时的默认值
    pre_save_hook: Optional[Callable[['models.Model', Dict[str, Any]], None]] = None  # 保存前钩子
    unique_field_case_insensitive: bool = True  # 唯一字段是否不区分大小写


def export_model(queryset, config: ExportConfig) -> Any:
    """
    通用导出函数

    Args:
        queryset: Django QuerySet
        config: ExportConfig 导出配置

    Returns:
        HttpResponse: Excel 文件响应
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse(
            'Excel 导出功能需要安装 openpyxl 库。请运行: pip install openpyxl',
            status=500,
            content_type='text/plain; charset=utf-8'
        )

    from django.utils import timezone
    from .export_utils import create_excel_response

    filename = f'{config.filename}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = config.sheet_title

    headers = [f.header for f in config.fields]

    # 样式函数
    def style_header(cell):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def style_data(cell):
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        style_header(cell)

    # 写入数据
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(config.fields, 1):
            value = field.getter(obj)
            if value is None:
                value = ''
            elif hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif hasattr(value, 'username'):
                value = value.username
            elif hasattr(value, 'name'):
                value = value.name
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
            style_data(ws.cell(row=row_num, column=col_num))

    # 调整列宽
    if config.column_widths:
        for col_num, width in enumerate(config.column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 创建响应
    response = create_excel_response(filename)
    wb.save(response)
    return response


def import_model(file, config: ImportConfig, user=None) -> Dict[str, Any]:
    """
    通用导入函数

    Args:
        file: 上传的 Excel 文件
        config: ImportConfig 导入配置
        user: 当前用户（用于审计）

    Returns:
        dict: 包含 success_count, created_count, updated_count, error_count, errors 的字典
    """
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        return {
            'success_count': 0,
            'error_count': 1,
            'errors': ['Excel 导入功能需要安装 openpyxl 库。请运行: pip install openpyxl']
        }

    try:
        import io
        from openpyxl.utils.exceptions import InvalidFileException

        file_content = file.read()
        if not file_content:
            return {
                'success_count': 0,
                'error_count': 1,
                'errors': ['文件内容为空']
            }

        try:
            wb = load_workbook(io.BytesIO(file_content))
        except Exception:
            wb = Workbook(io.BytesIO(file_content))

        ws = wb.active
        if ws is None:
            return {
                'success_count': 0,
                'error_count': 1,
                'errors': ['Excel 文件格式不正确，无法读取工作表']
            }

        if ws.max_row < 1:
            return {
                'success_count': 0,
                'error_count': 1,
                'errors': ['Excel 文件为空或格式不正确']
            }

        first_row = ws[1]
        if first_row is None:
            return {
                'success_count': 0,
                'error_count': 1,
                'errors': ['Excel 文件无法读取第一行']
            }

        headers = [cell.value for cell in first_row]

        # 建立列索引映射
        col_map = {}  # {field_name: col_index}
        for idx, h in enumerate(headers, 1):
            if not h:
                continue
            h_lower = str(h).strip().lower()
            for field in config.field_mappings:
                if h_lower in [alt.lower() for alt in field.excel_headers]:
                    col_map[field.model_field] = idx
                    break

        # 检查必填字段
        required_fields = {config.unique_field}
        for field in config.field_mappings:
            if field.model_field not in col_map:
                if field.model_field in required_fields:
                    return {
                        'success_count': 0,
                        'error_count': 1,
                        'errors': [f'未找到必填字段: {field.excel_headers[0]}']
                    }

        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        imported_keys = set()  # 跟踪本批次内的唯一值，用于检测重复行

        row_iterator = ws.iter_rows(min_row=2, values_only=True)
        for row_num, row in enumerate(row_iterator, 2):
            try:
                if row is None:
                    continue

                data = {}
                unique_value = None

                # 提取各字段值
                for field in config.field_mappings:
                    col_idx = col_map.get(field.model_field)
                    if col_idx and col_idx <= len(row):
                        raw_val = row[col_idx - 1]
                        if raw_val is not None:
                            try:
                                data[field.model_field] = field.transformer(raw_val)
                            except Exception:
                                data[field.model_field] = field.transformer(str(raw_val))
                        else:
                            data[field.model_field] = None

                    # 获取唯一字段值
                    if field.model_field == config.unique_field:
                        raw_val = row[col_idx - 1] if col_idx else None
                        if raw_val is not None:
                            unique_value = str(raw_val).strip()

                # 验证唯一字段
                if not unique_value:
                    errors.append(f'第{row_num}行: {config.unique_field}不能为空')
                    error_count += 1
                    continue

                # 检查批次内重复
                unique_key = unique_value.lower() if config.unique_field_case_insensitive else unique_value
                if unique_key in imported_keys:
                    errors.append(f'第{row_num}行: {config.unique_field} "{unique_value}" 在本次导入中已出现，将跳过')
                    error_count += 1
                    continue
                imported_keys.add(unique_key)

                # 查询现有记录
                filter_kwargs = {f'{config.unique_field}__iexact': unique_value} if config.unique_field_case_insensitive else {config.unique_field: unique_value}
                existing = config.model.objects.filter(**filter_kwargs).first()

                if existing:
                    # 更新现有记录
                    for field_name in config.update_fields:
                        if field_name in data and data[field_name] is not None:
                            setattr(existing, field_name, data[field_name])
                    if config.pre_save_hook:
                        config.pre_save_hook(existing, data)
                    existing.save()
                    updated_count += 1
                else:
                    # 创建新记录
                    create_data = {}
                    for field_name, default_val in config.create_defaults.items():
                        if callable(default_val):
                            create_data[field_name] = default_val(user)
                        else:
                            create_data[field_name] = default_val
                    for field in config.field_mappings:
                        if field.model_field in data and data[field.model_field] is not None:
                            create_data[field.model_field] = data[field.model_field]
                    if config.pre_save_hook:
                        config.pre_save_hook(None, create_data)
                    config.model.objects.create(**create_data)
                    created_count += 1

            except Exception as e:
                errors.append(f'第{row_num}行: {str(e)}')
                error_count += 1

        return {
            'success_count': created_count + updated_count,
            'created_count': created_count,
            'updated_count': updated_count,
            'error_count': error_count,
            'errors': errors[:50]
        }

    except InvalidFileException as e:
        return {
            'success_count': 0,
            'error_count': 1,
            'errors': [f'无效的 Excel 文件: {str(e)}']
        }
    except Exception as e:
        return {
            'success_count': 0,
            'error_count': 1,
            'errors': [f'文件读取失败: {str(e)}']
        }
